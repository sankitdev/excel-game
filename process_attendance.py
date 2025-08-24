import sys
from pathlib import Path
from openpyxl import load_workbook
import os

def is_missing(v):
    """Check if a value is considered missing (None, empty, or '-')."""
    return v is None or (isinstance(v, str) and v.strip() in ("", "-"))

def find_header_row_and_cols(ws, search_rows=20, max_cols=50):
    """
    Scan the top of the sheet to find the header row containing:
    - Name or Employee Name
    - Check In Time
    - Check Out Time
    Returns (header_row_index, name_col, checkin_col, checkout_col)
    """
    name_headers = {"name", "employee name"}
    cin_headers = {"check in time", "checkin time", "check in", "checkin"}
    cout_headers = {"check out time", "checkout time", "check out", "checkout"}

    for r in range(1, min(search_rows, ws.max_row) + 1):
        labels = {}
        for c in range(1, min(max_cols, ws.max_column) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str):
                val = v.strip().lower()
                labels[val] = c
        # find required columns in this row
        name_col = next((labels[k] for k in name_headers if k in labels), None)
        cin_col = next((labels[k] for k in cin_headers if k in labels), None)
        cout_col = next((labels[k] for k in cout_headers if k in labels), None)

        if name_col and cin_col and cout_col:
            return r, name_col, cin_col, cout_col

    raise RuntimeError("Could not find header row with Name / Check In Time / Check Out Time")

def process_ws(ws, progress_callback=None):
    # 1) Detect header and column indices
    header_row, name_col, cin_col, cout_col = find_header_row_and_cols(ws)
    shift_col = 4  # Assuming Shift is the 4th column (1-based index) based on provided data
    timetable_col = 5  # Assuming Timetable is the 5th column (1-based index)

    # 2) Iterate from last data row to first, marking rows for deletion
    first_data_row = header_row + 1
    last_row = ws.max_row
    rows_to_delete = []
    total_rows = last_row - first_data_row + 1
    processed_rows = 0

    r = last_row
    while r >= first_data_row:
        name = ws.cell(r, name_col).value
        cin_val = ws.cell(r, cin_col).value
        cout_val = ws.cell(r, cout_col).value
        shift_val = ws.cell(r, shift_col).value
        timetable_val = ws.cell(r, timetable_col).value

        # Check if both check-in and check-out are missing and a shift is scheduled
        if is_missing(cin_val) and is_missing(cout_val) and shift_val and timetable_val and "-" not in (shift_val, timetable_val):
            rows_to_delete.append(r)

        r -= 1
        processed_rows += 1
        if progress_callback:
            progress = int((processed_rows / total_rows) * 20)  # 20% for deletion phase
            progress_callback(progress)  # // grok show progress

   # 3) Delete marked rows in batches for efficiency (group consecutive rows)
    if rows_to_delete:
        rows_to_delete.sort()  # Sort ascending to group ranges
        batch_starts = [rows_to_delete[0]]
        batch_lengths = [1]
        for i in range(1, len(rows_to_delete)):
            if rows_to_delete[i] == rows_to_delete[i-1] + 1:
                batch_lengths[-1] += 1
            else:
                batch_starts.append(rows_to_delete[i])
                batch_lengths.append(1)
        
        # Delete from bottom to top to avoid index shifts
        for start, length in zip(reversed(batch_starts), reversed(batch_lengths)):
            ws.delete_rows(start, length)
        
        if progress_callback:
            progress = 20
            progress_callback(progress)  # // grok show progress

    # 4) Process blocks of the same employee for filling missing values
    r = first_data_row
    total_blocks = 0
    current_block = 0

    # Count total blocks for progress tracking
    while r <= ws.max_row:
        name = ws.cell(r, name_col).value
        start = r
        while r <= ws.max_row and ws.cell(r, name_col).value == name:
            r += 1
        end = r - 1
        if start <= end:
            total_blocks += 1
        r = end + 1
    r = first_data_row  # Reset for actual processing

    while r <= ws.max_row:
        name = ws.cell(r, name_col).value

        # Define the start and end of the current employee's block
        start = r
        while r <= ws.max_row and ws.cell(r, name_col).value == name:
            r += 1
        end = r - 1  # inclusive
        current_block += 1

        # 5) Backward fill: Look upward for missing values
        for i in range(start, end + 1):
            cin_cell = ws.cell(i, cin_col)
            cout_cell = ws.cell(i, cout_col)
            cin_val = cin_cell.value
            cout_val = cout_cell.value

            # Skip if both missing (already handled by deletion)
            if is_missing(cin_val) and is_missing(cout_val):
                continue

            need_cin = is_missing(cin_val) and not is_missing(cout_val)
            need_cout = is_missing(cout_val) and not is_missing(cin_val)

            if not need_cin and not need_cout:
                continue

            # Search upward within the block for the closest previous value
            j = i - 1
            while j >= start and ws.cell(j, name_col).value == name and (need_cin or need_cout):
                prev_cin = ws.cell(j, cin_col).value
                prev_cout = ws.cell(j, cout_col).value

                if need_cin and not is_missing(prev_cin):
                    cin_cell.value = prev_cin
                    need_cin = False

                if need_cout and not is_missing(prev_cout):
                    cout_cell.value = prev_cout
                    need_cout = False

                j -= 1

        # 6) Forward fill: Look downward for remaining missing values
        for i in range(end, start - 1, -1):  # Iterate from bottom to top
            cin_cell = ws.cell(i, cin_col)
            cout_cell = ws.cell(i, cout_col)
            cin_val = cin_cell.value
            cout_val = cout_cell.value

            # Skip if both missing or if already filled
            if is_missing(cin_val) and is_missing(cout_val):
                continue

            need_cin = is_missing(cin_val) and not is_missing(cout_val)
            need_cout = is_missing(cout_val) and not is_missing(cin_val)

            if not need_cin and not need_cout:
                continue

            # Search downward within the block for the closest next value
            j = i + 1
            while j <= end and ws.cell(j, name_col).value == name and (need_cin or need_cout):
                next_cin = ws.cell(j, cin_col).value
                next_cout = ws.cell(j, cout_col).value

                if need_cin and not is_missing(next_cin):
                    cin_cell.value = next_cin
                    need_cin = False

                if need_cout and not is_missing(next_cout):
                    cout_cell.value = next_cout
                    need_cout = False

                j += 1

        if progress_callback:
            progress = 20 + int((current_block / total_blocks) * 60)  # 60% for fill phases
            progress_callback(progress)  # // grok show progress

    if progress_callback:
        progress = 80  # 80% after filling, leaving 20% for saving
        progress_callback(progress)  # // grok show progress

def process_excel(file, progress_callback=None):
    """
    file: either a file path (str/Path) or file-like object (BytesIO / Streamlit UploadedFile)
    Returns: openpyxl Workbook object
    """
    # detect if .xlsm
    name = getattr(file, "name", "")  # UploadedFile has .name
    suffix = Path(name).suffix.lower() if name else ""
    keep_vba = suffix == ".xlsm"

    wb = load_workbook(file, keep_vba=keep_vba)
    ws = wb.worksheets[0]  # first sheet only
    process_ws(ws, progress_callback)
    return wb

def main():
    if len(sys.argv) < 2:
        print("Usage: python process_attendance.py <input.xlsx>")
        sys.exit(1)

    inp = Path(sys.argv[1])
    if not inp.exists():
        print(f"Input file not found: {inp}")
        sys.exit(1)

    # Auto-generate output filename
    base, ext = os.path.splitext(str(inp))
    out = f"{base}_processed{ext}"

    # Load workbook safely (don't force keep_vba=True unless it's .xlsm)
    if inp.suffix.lower() == ".xlsm":
        wb = load_workbook(inp, keep_vba=True)
    else:
        wb = load_workbook(inp)

    # Process ONLY first sheet
    ws = wb.worksheets[0]
    process_ws(ws)

    wb.save(out)
    print(f"Done. Wrote: {out}")

if __name__ == "__main__":
    main()